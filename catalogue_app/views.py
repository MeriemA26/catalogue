# catalogue_app/views.py - Version avec corrections
from aiohttp import request
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import os
import json
import logging
import base64
from decimal import Decimal
from django.http import StreamingHttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import time
from .models import Enseigne, Catalogue, Produit
from .forms import UploadForm, ProduitForm
from .utils import OCRProcessor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from .sql_sync import SQLServerSync

sql_sync = SQLServerSync()
# Configuration du logging - Désactiver complètement
logging.disable(logging.CRITICAL)
logger = logging.getLogger(__name__)
# catalogue_app/views.py

def supprimer_catalogues_vides():
    """Supprime tous les catalogues qui n'ont plus de produits"""
    from django.db import models
    catalogues_vides = Catalogue.objects.annotate(
        nb_produits=models.Count('produits')
    ).filter(nb_produits=0)
    
    count = catalogues_vides.count()
    if count > 0:
        for c in catalogues_vides:
            print(f"🗑️ Suppression du catalogue vide ID {c.id} - {c.enseigne.nom}")
            c.delete()
        print(f"✅ {count} catalogues vides supprimés")
    return count

def index(request):
    """Page d'accueil - Tableau de bord simplifié"""
    from django.db.models import Count, Q
    from datetime import datetime
    supprimer_catalogues_vides()
    # Statistiques de base
    total_catalogues = Catalogue.objects.count()
    total_produits_sauvegardes = Produit.objects.filter(est_sauvegarde=True).count()
    total_enseignes = Enseigne.objects.count()
    
    # Produits à traiter (non sauvegardés)
    produits_a_traiter = Produit.objects.filter(est_sauvegarde=False).count()
    
    # Catalogues actifs avec comptage des produits sauvegardés
    catalogues_actifs = Catalogue.objects.filter(
        date_fin__gte=datetime.now().date()
    ).annotate(
        nb_produits=Count('produits'),
        nb_sauvegardes=Count('produits', filter=Q(produits__est_sauvegarde=True)),
        nb_non_sauvegardes=Count('produits', filter=Q(produits__est_sauvegarde=False))
    ).filter(
        nb_produits__gt=0
    ).order_by('-date_debut')
    
    context = {
        'total_catalogues': total_catalogues,
        'total_produits_sauvegardes': total_produits_sauvegardes,
        'total_enseignes': total_enseignes,
        'produits_a_traiter': produits_a_traiter,
        'catalogues_actifs': catalogues_actifs,
    }
    return render(request, 'index.html', context)

def upload(request):
    """Page d'upload avec traitement OCR pour une ou plusieurs images"""
    # Initialiser last_uploaded_image
    last_uploaded_image = request.session.get('last_uploaded_image', None)
    last_uploaded_images = request.session.get('last_uploaded_images', [])
    
    #  Récupérer le catalogue actuel depuis la session
    catalogue_id = request.session.get('active_catalogue_id', None)
    catalogue_actuel = None
    
    if catalogue_id:
        try:
            catalogue_actuel = Catalogue.objects.get(id=catalogue_id)
            print(f"✅ Catalogue récupéré depuis la session: ID {catalogue_actuel.id}")
        except Catalogue.DoesNotExist:
            catalogue_actuel = None
            if 'active_catalogue_id' in request.session:
                del request.session['active_catalogue_id']
    
    #  Si pas de catalogue dans la session, vérifier s'il y a des produits non sauvegardés
    if not catalogue_actuel:
        produits_non_sauvegardes_total = Produit.objects.filter(est_sauvegarde=False).count()
        
        if produits_non_sauvegardes_total > 0:
            dernier_catalogue = Catalogue.objects.filter(
                produits__est_sauvegarde=False
            ).distinct().order_by('-date_upload').first()
            
            if dernier_catalogue:
                catalogue_actuel = dernier_catalogue
                request.session['active_catalogue_id'] = dernier_catalogue.id
                print(f"Catalogue récupéré depuis la base: ID {dernier_catalogue.id}")
        else:
            # Si plus de produits, nettoyer la session
            if 'last_uploaded_images' in request.session:
                del request.session['last_uploaded_images']
            if 'last_uploaded_image' in request.session:
                del request.session['last_uploaded_image']
            if 'active_catalogue_id' in request.session:
                del request.session['active_catalogue_id']
            last_uploaded_images = []
            last_uploaded_image = None
            print("Session nettoyée (plus de produits)")
    
    # Si pas d'image dans last_uploaded_image mais des images dans last_uploaded_images
    if not last_uploaded_image and last_uploaded_images:
        last_uploaded_image = last_uploaded_images[0] if last_uploaded_images else None
    
    # Récupérer l'image du catalogue si elle n'est pas dans la session
    if not last_uploaded_image and catalogue_actuel and catalogue_actuel.image_path:
        last_uploaded_image = os.path.join(settings.MEDIA_URL, catalogue_actuel.image_path)
        request.session['last_uploaded_image'] = last_uploaded_image
        if not last_uploaded_images:
            last_uploaded_images = [last_uploaded_image]
            request.session['last_uploaded_images'] = last_uploaded_images
        print(f"Image récupérée depuis le catalogue ID {catalogue_actuel.id}")
    
    # DEBUG - Afficher l'état de la session
    print("=" * 80)
    print("🔍 UPLOAD VIEW - ÉTAT DE LA SESSION")
    print(f"📌 last_uploaded_image depuis session: {last_uploaded_image}")
    print(f"📌 last_uploaded_images: {last_uploaded_images}")
    print(f"📌 active_catalogue_id: {catalogue_id}")
    print(f"📌 catalogue_actuel: {catalogue_actuel.id if catalogue_actuel else 'None'}")
    print(f"📌 Toute la session: {dict(request.session)}")
    print("=" * 80)

    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                enseigne = form.cleaned_data['enseigne']
                date_debut = form.cleaned_data['date_debut']
                date_fin = form.cleaned_data['date_fin']
                note = form.cleaned_data.get('note', '')
                images = request.FILES.getlist('images')
                
                if not images:
                    messages.error(request, "Veuillez sélectionner au moins une image.")
                    return redirect('upload')
                # VALIDATION : date_debut doit être < date_fin
                if date_debut >= date_fin:
                    messages.error(request, "La date de début doit être antérieure à la date de fin.")
                    return redirect('upload')
                # Vérifier si le catalogue existe déjà
                catalogue, created = Catalogue.objects.get_or_create(
                    enseigne=enseigne,
                    date_debut=date_debut,
                    date_fin=date_fin,
                    defaults={'date_upload': timezone.now()
                              , 'note': note}
                )
                if not created and note and catalogue.note != note:
                    catalogue.note = note
                    catalogue.save()
                #Sauvegarder l'ID du catalogue dans la session
                request.session['active_catalogue_id'] = catalogue.id
                
                #Si le catalogue existe déjà, récupérer les images existantes
                existing_images = request.session.get('last_uploaded_images', [])
                
                total_produits = 0
                images_paths = []
                
                # Traiter chaque image
                for idx, image in enumerate(images):
                    # Sauvegarder l'image du catalogue
                    image_name = f'{enseigne.nom}_{date_debut}_{date_fin}_{idx}_{image.name}'
                    image_path = default_storage.save(
                        f'catalogues/{image_name}',
                        ContentFile(image.read())
                    )
                    images_paths.append(image_path)
                    
                    # Traitement OCR
                    ocr = OCRProcessor()
                    full_path = os.path.join(settings.MEDIA_ROOT, image_path)
                    
                    if not os.path.exists(full_path):
                        continue
                    
                    # Traiter l'image
                    produits_data = []
                    try:
                        produits_data = ocr.traiter_image_complete(full_path)
                        print(f"Image {idx+1}: {len(produits_data)} produits détectés")
                        
                        # Afficher les données extraites pour debug
                        print("\n" + "=" * 80)
                        print(f"DONNEES EXTRAITES PAR L'OCR - IMAGE {idx+1}:")
                        print("=" * 80)
                        for i, data in enumerate(produits_data):
                            print(f"\nProduit {i+1}:")
                            print(f"  Nom (FR): {data.get('nom', 'NON TROUVE')}")
                            print(f"  Nom (AR): {data.get('nom_ar', 'NON TROUVE')}")
                            print(f"  Marque: {data.get('marque', 'NON TROUVE')}")
                            print(f"  Prix: {data.get('prix', 'NON TROUVE')}")
                            print(f"  Prix Avant: {data.get('prix_avant', 'NON TROUVE')}")
                            print(f"  %: {data.get('pourcentage', 'NON TROUVE')}")
                            print(f"  Remise: {data.get('remise', 'NON TROUVE')}")
                            print(f"  Desc_1: {data.get('desc_1', 'NON TROUVE')[:50]}...")
                            print(f"  Desc_2: {data.get('desc_2', 'NON TROUVE')[:50]}...")
                            print(f"  Desc_3: {data.get('desc_3', 'NON TROUVE')[:50]}...")
                        print("=" * 80 + "\n")
                        
                    except Exception as e:
                        print(f"Erreur YOLO sur image {idx+1}: {e}")
                    
                    # Fallback si nécessaire
                    if not produits_data:
                        print(f"Utilisation du mode fallback (EasyOCR) pour l'image {idx+1}")
                        try:
                            texte = ocr.extraire_texte_fallback(full_path)
                            if texte:
                                produits_data = ocr.detecter_prix_remise(texte)
                                print(f"📊 Fallback image {idx+1}: {len(produits_data)} produits")
                        except Exception as e:
                            print(f"Erreur fallback: {e}")
                    
                    # Créer les produits dans la base de données
                    if produits_data:
                        for pidx, data in enumerate(produits_data):
                            try:
                                # Sauvegarder l'image du produit si disponible
                                image_produit = None
                                if 'product_image_b64' in data and data['product_image_b64']:
                                    try:
                                        image_data = base64.b64decode(data['product_image_b64'])
                                        timestamp = int(timezone.now().timestamp())
                                        filename = f'produit_{catalogue.id}_{idx}_{pidx}_{timestamp}.jpg'
                                        image_produit = ContentFile(image_data, filename)
                                        print(f"Image produit {pidx} decodee avec succes")
                                    except Exception as e:
                                        print(f"Erreur decodage image produit {pidx}: {e}")
                                
                                # Récupérer les noms exacts
                                nom_fr = data.get('nom_fr', '').strip()
                                nom_ar = data.get('nom_ar', '').strip()
                                
                                if nom_fr:
                                    nom_affiche = nom_fr
                                elif nom_ar:
                                    nom_affiche = nom_ar
                                else:
                                    nom_affiche = f"Produit {idx}_{pidx + 1}"
                                
                                # Nettoyer les descriptions
                                desc_1= data.get('desc_1', '') or ''
                                desc_2 = data.get('desc_2', '') or ''
                                desc_3 = data.get('desc_3', '') or ''
                                
                                if not desc_1 and (desc_2 or desc_3):
                                    desc_1 = desc_2
                                    desc_2 = desc_3
                                    desc_3 = ''
                                
                                remise_val = data.get('remise')
                                if isinstance(remise_val, str) and '%' in remise_val:
                                    remise_val = None
                                
                                produit = Produit(
                                    catalogue=catalogue,
                                    nom=nom_affiche,
                                    nom_fr=nom_fr if nom_fr else None,
                                    nom_ar=nom_ar if nom_ar else None,
                                    marque=data.get('marque', '') if data.get('marque') else None,
                                    prix=data.get('prix'),
                                    prix_avant=data.get('prix_avant'),
                                    pourcentage=data.get('pourcentage'),
                                    remise=remise_val if remise_val else None,
                                    desc_1=desc_1 if desc_1 else None,
                                    desc_2=desc_2 if desc_2 else None,
                                    desc_3=desc_3 if desc_3 else None,
                                    extrait_texte=data.get('extrait_texte', '') if data.get('extrait_texte') else None,
                                    image_produit=image_produit,
                                )
                                produit.save()
                                total_produits += 1
                                print(f"Produit cree: ID {produit.id}")
                                
                            except Exception as e:
                                print(f"Erreur creation produit {idx}_{pidx}: {e}")
                                import traceback
                                traceback.print_exc()
                                continue
                
                # 🔥 Sauvegarder TOUTES les images dans la session (conserver les anciennes + nouvelles)
                all_image_urls = []
                
                # 1. Ajouter les images existantes
                if existing_images:
                    all_image_urls.extend(existing_images)
                    print(f"📌 {len(existing_images)} images existantes conservées")
                
                # 2. Ajouter les nouvelles images
                for path in images_paths:
                    image_url = os.path.join(settings.MEDIA_URL, path)
                    # Éviter les doublons
                    if image_url not in all_image_urls:
                        all_image_urls.append(image_url)
                
                if all_image_urls:
                    request.session['last_uploaded_images'] = all_image_urls
                    request.session['last_uploaded_image'] = all_image_urls[0]
                    last_uploaded_image = all_image_urls[0]
                    last_uploaded_images = all_image_urls
                    
                    # Sauvegarder aussi le chemin de la première image dans le catalogue
                    if not catalogue.image_path:
                        catalogue.image_path = images_paths[0] if images_paths else None
                        catalogue.save()
                    
                    print(f"✅ {len(all_image_urls)} images totales sauvegardées dans la session")
                
                if total_produits > 0:
                    messages.success(request, f'Upload réussi ! {total_produits} nouveaux produits détectés sur {len(images)} page(s).')
                else:
                    messages.warning(request, f'Aucun nouveau produit détecté dans les {len(images)} images.')
                
            except Exception as e:
                print(f"Erreur generale: {e}")
                messages.error(request, f"Erreur lors du traitement: {str(e)}")
                import traceback
                traceback.print_exc()
            
            return redirect('upload')
    else:
        # 🔥 Pré-remplir le formulaire avec les valeurs du catalogue actuel
        form = UploadForm()
        
        if catalogue_actuel:
            initial_data = {
                'enseigne': catalogue_actuel.enseigne.id,
                'date_debut': catalogue_actuel.date_debut.strftime('%Y-%m-%d'),
                'date_fin': catalogue_actuel.date_fin.strftime('%Y-%m-%d'),
                'note': catalogue_actuel.note or '',  # 🔥 Pré-remplir la note
            }
            form = UploadForm(initial=initial_data)
            print(f"✅ Formulaire pré-rempli avec les valeurs du catalogue {catalogue_actuel.id}")
            print(f"   Note: '{catalogue_actuel.note}'")

    # 🔥 Récupérer TOUS les produits non sauvegardés dans UN SEUL tableau
    catalogues_recents = {}
    produits_non_sauvegardes = 0
    
    # Récupérer tous les produits non sauvegardés
    tous_produits = Produit.objects.filter(est_sauvegarde=False).order_by('-created_at')
    
    if tous_produits.exists():
        # 🔥 Utiliser le premier catalogue comme référence
        premier_produit = tous_produits.first()
        if premier_produit and premier_produit.catalogue:
            catalogue_ref = premier_produit.catalogue
            enseigne_nom = catalogue_ref.enseigne.nom if catalogue_ref.enseigne else "Sans enseigne"
            key = f"{enseigne_nom} - {catalogue_ref.date_debut} au {catalogue_ref.date_fin}"
            
            # 🔥 Prendre tous les produits non sauvegardés (même de plusieurs catalogues)
            # mais les regrouper dans un seul tableau
            catalogues_recents[key] = tous_produits
            produits_non_sauvegardes = tous_produits.count()
            print(f"✅ {key}: {produits_non_sauvegardes} produits (regroupés)")
        else:
            # Fallback: regrouper par catalogue
            catalogues_avec_produits = Catalogue.objects.filter(
                produits__est_sauvegarde=False
            ).distinct().order_by('-date_upload')
            
            for catalogue in catalogues_avec_produits:
                enseigne_nom = catalogue.enseigne.nom if catalogue.enseigne else "Sans enseigne"
                produits_non_sauvegardes_catalogue = catalogue.produits.filter(est_sauvegarde=False)
                count = produits_non_sauvegardes_catalogue.count()
                
                if count > 0:
                    key = f"{enseigne_nom} - {catalogue.date_debut} au {catalogue.date_fin}"
                    catalogues_recents[key] = produits_non_sauvegardes_catalogue
                    produits_non_sauvegardes += count
                    print(f"✅ {key}: {count} produits")
    
    # 🔍 DEBUG - Afficher les produits trouvés
    print("=" * 80)
    print("🔍 UPLOAD VIEW - PRODUITS TROUVÉS")
    print(f"📌 catalogues_recents: {list(catalogues_recents.keys())}")
    print(f"📌 produits_non_sauvegardes: {produits_non_sauvegardes}")
    print(f"📌 last_uploaded_image final: {last_uploaded_image}")
    print(f"📌 last_uploaded_images: {last_uploaded_images}")
    print(f"📌 catalogue_actuel final: {catalogue_actuel.id if catalogue_actuel else 'None'}")
    print("=" * 80)

    total_produits = Produit.objects.count()
    produits_sauvegardes = Produit.objects.filter(est_sauvegarde=True).count()
    produits_non_sauvegardes_total = Produit.objects.filter(est_sauvegarde=False).count()
    
    context = {
        'form': form,
        'catalogues_recents': catalogues_recents,
        'enseignes': Enseigne.objects.all(),
        'total_catalogues': Catalogue.objects.count(),
        'total_produits': total_produits,
        'produits_sauvegardes': produits_sauvegardes,
        'produits_non_sauvegardes': produits_non_sauvegardes_total,
        'last_uploaded_image': last_uploaded_image,
        'last_uploaded_images': last_uploaded_images,
        'catalogue_actuel': catalogue_actuel,
    }
    return render(request, 'upload.html', context)

def upload_stream(request):
    """
    Vue SSE : reçoit les images uploadées, lance le pipeline produit par produit,
    et envoie chaque produit au client dès qu'il est prêt via Server-Sent Events.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requis'}, status=405)

    print("=" * 80)
    print("🔍 UPLOAD STREAM - DEBUG")
    print(f"📌 FILES: {request.FILES}")
    print(f"📌 POST: {request.POST}")
    print("=" * 80)

    # 🔥 Récupérer directement les fichiers
    images = request.FILES.getlist('images')
    
    if not images:
        return JsonResponse({'error': 'Aucune image sélectionnée'}, status=400)

    # Récupérer les autres données du POST
    enseigne_id = request.POST.get('enseigne')
    date_debut_str = request.POST.get('date_debut')
    date_fin_str = request.POST.get('date_fin')
    note = request.POST.get('note', '')
    if not enseigne_id or not date_debut_str or not date_fin_str:
        return JsonResponse({'error': 'Données manquantes'}, status=400)

    # Récupérer l'enseigne
    try:
        enseigne = Enseigne.objects.get(id=enseigne_id)
    except Enseigne.DoesNotExist:
        return JsonResponse({'error': 'Enseigne non trouvée'}, status=400)

    # Convertir les dates
    from datetime import datetime
    date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
    date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
    if date_debut >= date_fin:
        return JsonResponse({'error': 'La date de début doit être antérieure à la date de fin.'}, status=400)    
    try:
        catalogue = Catalogue.objects.get(
            enseigne=enseigne,
            date_debut=date_debut,
            date_fin=date_fin
        )
        # 🔥 Mettre à jour la note si elle a changé
        if note and catalogue.note != note:
            catalogue.note = note
            catalogue.save()
            print(f"✅ Note mise à jour dans upload_stream: '{note}'")
    except Catalogue.DoesNotExist:
        # Créer un nouveau catalogue avec la note
        catalogue = Catalogue.objects.create(
            enseigne=enseigne,
            date_debut=date_debut,
            date_fin=date_fin,
            date_upload=timezone.now(),
            note=note
        )
        print(f"✅ Nouveau catalogue créé avec note: '{note}'")

    total_produits = 0
    images_paths = []
    all_image_urls = []
    
    # 🔥 Récupérer les images existantes AVANT le traitement
    existing_images = request.session.get('last_uploaded_images', [])
    print(f"📌 Images existantes dans la session: {len(existing_images)}")

    def event_stream():
        nonlocal total_produits, images_paths, all_image_urls
        ocr = OCRProcessor()

        try:
            from .pipeline import (
                get_product_model, get_field_model, get_ocr_readers,
                FIELD_CLASS_MAP, image_to_base64,
                extract_price, extract_percentage, extract_text,
                _detect_language
            )
            import cv2, torch, re, base64 as b64
            from decimal import Decimal

            device = 'cuda' if torch.cuda.is_available() else 'cpu'
            product_model = get_product_model()
            field_model   = get_field_model()
            reader_latin, reader_ar = get_ocr_readers()

            # 🔥 Traiter CHAQUE image
            for img_idx, image in enumerate(images):
                # Sauvegarder l'image catalogue
                image_name = f'{enseigne.nom}_{date_debut}_{date_fin}_{img_idx}_{image.name}'
                image_path = default_storage.save(
                    f'catalogues/{image_name}',
                    ContentFile(image.read())
                )
                images_paths.append(image_path)
                image_url = os.path.join(settings.MEDIA_URL, image_path)
                all_image_urls.append(image_url)
                
                # Si c'est la première image, la sauvegarder dans le catalogue
                if img_idx == 0:
                    catalogue.image_path = image_path
                    catalogue.save()
                    media_image_url = os.path.join(settings.MEDIA_URL, image_path)
                    yield f"data: {json.dumps({'type': 'catalogue', 'image_url': media_image_url, 'catalogue_id': catalogue.id, 'page': img_idx + 1, 'total_pages': len(images)})}\n\n"

                full_path = os.path.join(settings.MEDIA_ROOT, image_path)

                image_cv = cv2.imread(full_path)
                if image_cv is None:
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Impossible de lire l image {img_idx + 1}'})}\n\n"
                    continue

                product_results = product_model(full_path, conf=0.5, verbose=False, device=device)[0]
                total = len(product_results.boxes)
                total_produits += total

                # Envoyer le nombre de produits pour cette image
                yield f"data: {json.dumps({'type': 'page_total', 'page': img_idx + 1, 'count': total, 'total_pages': len(images)})}\n\n"

                if total == 0:
                    yield f"data: {json.dumps({'type': 'page_done', 'page': img_idx + 1, 'count': 0})}\n\n"
                    continue

                for idx, box in enumerate(product_results.boxes):
                    x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                    crop = image_cv[y1:y2, x1:x2]
                    if crop.size == 0:
                        continue

                    data = {
                        'nom_fr': '', 'nom_ar': '', 'marque': '',
                        'prix': None, 'prix_avant': None, 'pourcentage': None,
                        'remise': '', 'desc_1': '', 'desc_2': '', 'desc_3': '',
                        'product_image_b64': image_to_base64(crop),
                    }

                    field_results = field_model.predict(crop, conf=0.5, device=device, verbose=False)[0]
                    field_names   = field_results.names
                    extracted_pct = None
                    extracted_prix = None
                    extracted_prix_avant = None

                    for fbox, fcls in zip(field_results.boxes.xyxy.cpu().numpy(), field_results.boxes.cls.cpu().numpy()):
                        fx1, fy1, fx2, fy2 = map(int, fbox)
                        class_name = field_names[int(fcls)]
                        roi = crop[fy1:fy2, fx1:fx2]
                        if roi.size == 0:
                            continue
                        target = FIELD_CLASS_MAP.get(class_name)
                        if target is None:
                            continue

                        if class_name == 'product_AR':
                            data['nom_ar'] = extract_text(roi, reader_ar)
                        elif class_name == 'product_name':
                            data['nom_fr'] = extract_text(roi, reader_latin)
                        elif class_name == 'brand':
                            v = extract_text(roi, reader_latin)
                            data['marque'] = v or extract_text(roi, reader_ar)
                        elif class_name == 'price':
                            v = extract_price(roi)
                            print(f"  🔍 extract_price retourné: '{v}'")
                            if v:
                                data['prix'] = Decimal(v)
                                extracted_prix = Decimal(v)
                                print(f"  ✅ Prix extrait: {v} -> {data['prix']}")
                        elif class_name == 'price_before':
                            v = extract_price(roi)
                            print(f"  🔍 extract_price (price_before) retourné: '{v}'")
                            if v:
                                data['prix_avant'] = Decimal(v)
                                extracted_prix_avant = Decimal(v)
                                print(f"  ✅ Prix avant extrait: {v} -> {data['prix_avant']}")
                        elif class_name == 'pct':
                            v = extract_percentage(roi)
                            print(f"  🔍 extract_percentage retourné: '{v}'")
                            if v:
                                data['pourcentage'] = Decimal(v)
                                extracted_pct = Decimal(v)
                                print(f"  ✅ Pourcentage extrait: {v} -> {data['pourcentage']}")
                        # ✅ CORRIGÉ :
                        elif class_name in ('description', 'description2', 'description3'):
                            v = extract_text(roi, reader_latin) or extract_text(roi, reader_ar)
                            if v:
                                v = re.sub(r'[^\w\s\u0600-\u06FF\.\,\-\(\)\:]+', ' ', v)
                                v = re.sub(r'\s+', ' ', v).strip()
                            if class_name == 'description':
                                data['desc_1'] = v or ''
                            elif class_name == 'description2':
                                data['desc_2'] = v or ''
                            elif class_name == 'description3':
                                data['desc_3'] = v or ''

                    # 🔥 LOGS DES VALEURS EXTRAITES
                    print(f"\n🔍 PRODUIT {idx+1} - VALEURS EXTRAITES BRUTES:")
                    print(f"  raw prix: {data.get('prix')}")
                    print(f"  raw prix_avant: {data.get('prix_avant')}")
                    print(f"  raw pourcentage: {data.get('pourcentage')}")
                    print(f"  extracted_prix: {extracted_prix}")
                    print(f"  extracted_prix_avant: {extracted_prix_avant}")
                    print(f"  extracted_pct: {extracted_pct}")

                    # 🔥 CORRECTION : Calcul des prix avec Decimal
                    if extracted_pct is not None and 1 <= float(extracted_pct) <= 100:
                        data['pourcentage'] = extracted_pct
                        data['remise'] = f"{extracted_pct}%"
                        
                        if extracted_prix is not None and extracted_prix > 0:
                            # prix_avant = prix / (1 - pct/100)
                            prix_avant_calc = extracted_prix / (1 - extracted_pct / 100)
                            data['prix_avant'] = round(prix_avant_calc, 3)
                            data['prix'] = extracted_prix
                            print(f"  ✅ Calcul CAS 1: prix={extracted_prix}, prix_avant={data['prix_avant']}, pct={extracted_pct}")
                        elif extracted_prix_avant is not None and extracted_prix_avant > 0:
                            # prix = prix_avant * (1 - pct/100)
                            prix_calc = extracted_prix_avant * (1 - extracted_pct / 100)
                            data['prix'] = round(prix_calc, 3)
                            data['prix_avant'] = extracted_prix_avant
                            print(f"  ✅ Calcul CAS 2: prix={data['prix']}, prix_avant={extracted_prix_avant}, pct={extracted_pct}")
                        else:
                            print(f"  ⚠️ Pourcentage trouvé mais pas de prix valide")
                    
                    elif extracted_prix is not None and extracted_prix_avant is not None and extracted_prix_avant > 0:
                        if extracted_prix < extracted_prix_avant:
                            pct = (1 - extracted_prix / extracted_prix_avant) * 100
                            if 1 <= float(pct) <= 100:
                                data['pourcentage'] = round(pct, 2)
                                data['remise'] = f"{round(pct, 2)}%"
                                data['prix'] = extracted_prix
                                data['prix_avant'] = extracted_prix_avant
                                print(f"  ✅ Calcul CAS 3: prix={extracted_prix}, prix_avant={extracted_prix_avant}, pct={round(pct, 2)}")
                            else:
                                print(f"  ⚠️ Pourcentage hors limite: {pct}")
                        else:
                            print(f"  ⚠️ Prix >= prix_avant: {extracted_prix} >= {extracted_prix_avant}")
                    elif extracted_prix is not None and extracted_prix > 0:
                        data['prix'] = extracted_prix
                        print(f"  ✅ Prix seul: {extracted_prix}")
                    elif extracted_prix_avant is not None and extracted_prix_avant > 0:
                        data['prix_avant'] = extracted_prix_avant
                        print(f"  ✅ Prix avant seul: {extracted_prix_avant}")
                    else:
                        print(f"  ⚠️ Aucun prix valide détecté")

                    # Créer le produit en DB
                    nom_fr = data['nom_fr'].strip()
                    nom_ar = data['nom_ar'].strip()
                    nom_affiche = nom_fr or nom_ar or f"Produit {idx + 1}"

                    desc_1   = data.get('desc_1', '') or ''
                    desc_2 = data.get('desc_2', '') or ''
                    desc_3 = data.get('desc_3', '') or ''
                    if not desc_1 and (desc_2 or desc_3):
                        desc_1, desc_2, desc_3 = desc_2, desc_3, ''

                    remise_val = data.get('remise')
                    if isinstance(remise_val, str) and '%' in remise_val:
                        remise_val = None

                    image_produit = None
                    if data.get('product_image_b64'):
                        try:
                            img_data = b64.b64decode(data['product_image_b64'])
                            ts = int(timezone.now().timestamp())
                            image_produit = ContentFile(img_data, f'produit_{catalogue.id}_{img_idx}_{idx}_{ts}.jpg')
                        except Exception:
                            pass

                    print(f"\n{'='*60}")
                    print(f"📦 CRÉATION PRODUIT {idx+1} - VALEURS DATA")
                    print(f"{'='*60}")
                    print(f"  data['prix']: {data.get('prix')} (type: {type(data.get('prix'))})")
                    print(f"  data['prix_avant']: {data.get('prix_avant')} (type: {type(data.get('prix_avant'))})")
                    print(f"  data['pourcentage']: {data.get('pourcentage')} (type: {type(data.get('pourcentage'))})")
                    print(f"  data['remise']: {data.get('remise')}")
                    print(f"  data['desc_1']: {data.get('desc_1')[:50] if data.get('desc_1') else 'None'}")
                    print(f"{'='*60}")

                    # 🔥 Créer le produit avec Decimal pour prix et prix_avant
                    produit = Produit(
                        catalogue=catalogue,
                        nom=nom_affiche,
                        nom_fr=nom_fr or None,
                        nom_ar=nom_ar or None,
                        marque=data.get('marque') or None,
                        prix=float(data.get('prix')) if data.get('prix') is not None else None,
                        prix_avant=float(data.get('prix_avant')) if data.get('prix_avant') is not None else None,
                        pourcentage=float(data.get('pourcentage')) if data.get('pourcentage') is not None else None,
                        remise=remise_val or None,
                        desc_1=desc_1 or None,
                        desc_2=desc_2 or None,
                        desc_3=desc_3 or None,
                        extrait_texte=f"Marque: {data.get('marque','')} - {desc_1}" or None,
                        image_produit=image_produit,
                    )
                    
                    print(f"\n  📊 VALEURS PRODUIT AVANT SAVE:")
                    print(f"  produit.prix: {produit.prix}")
                    print(f"  produit.prix_avant: {produit.prix_avant}")
                    print(f"  produit.pourcentage: {produit.pourcentage}")
                    print(f"  produit.remise: {produit.remise}")
                    
                    produit.save()
                    
                    print(f"\n  📊 VALEURS PRODUIT APRÈS SAVE:")
                    print(f"  produit.prix: {produit.prix}")
                    print(f"  produit.prix_avant: {produit.prix_avant}")
                    print(f"  produit.pourcentage: {produit.pourcentage}")
                    print(f"  produit.remise: {produit.remise}")
                    print(f"{'='*60}\n")

                    # 🔥 Préparer la payload SSE (UNE SEULE FOIS)
                    prix_val = str(produit.prix) if produit.prix is not None else ''
                    prix_avant_val = str(produit.prix_avant) if produit.prix_avant is not None else ''
                    pct_val = str(int(produit.pourcentage)) if produit.pourcentage is not None else ''
                    
                    # 🔥 S'assurer que le prix a 3 décimales
                    if prix_val and '.' in prix_val:
                        parts = prix_val.split('.')
                        if len(parts[1]) < 3:
                            prix_val = parts[0] + '.' + parts[1].ljust(3, '0')
                    elif prix_val:
                        prix_val = prix_val + '.000'

                    print(f"\n  📤 PAYLOAD SSE - Produit {idx+1}:")
                    print(f"  prix: '{prix_val}'")
                    print(f"  prix_avant: '{prix_avant_val}'")
                    print(f"  pourcentage: '{pct_val}'")

                    # 🔥 UNE SEULE PAYLOAD
                    payload = {
                        'type': 'product',
                        'index': idx + 1,
                        'page': img_idx + 1,
                        'id': produit.id,
                        'nom': nom_affiche,
                        'nom_fr': nom_fr,
                        'nom_ar': nom_ar,
                        'marque': data.get('marque', ''),
                        'prix': prix_val,
                        'prix_avant': prix_avant_val,
                        'pourcentage': pct_val,
                        'desc_1': desc_1[:80] if desc_1 else '',
                        'desc_2': desc_2[:80] if desc_2 else '',
                        'desc_3': desc_3[:80] if desc_3 else '',
                        'image_url': produit.image_produit.url if produit.image_produit else '',
                    }
                    yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

                yield f"data: {json.dumps({'type': 'page_done', 'page': img_idx + 1, 'count': total})}\n\n"

            # 🔥 Combiner les images existantes et nouvelles
            final_image_urls = []
            if existing_images:
                final_image_urls.extend(existing_images)
                print(f"📌 {len(existing_images)} images existantes conservées")
            
            for url in all_image_urls:
                if url not in final_image_urls:
                    final_image_urls.append(url)
            
            if final_image_urls:
                request.session['last_uploaded_images'] = final_image_urls
                request.session['last_uploaded_image'] = final_image_urls[0]
                request.session.modified = True
                request.session.save()
                print(f"✅ {len(final_image_urls)} images totales sauvegardées dans la session")

            # Fin du traitement - toutes les images traitées
            yield f"data: {json.dumps({'type': 'done', 'count': total_produits, 'pages': len(images), 'images': final_image_urls})}\n\n"

        except Exception as e:
            import traceback
            traceback.print_exc()
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response
# catalogue_app/views.py - edit_product CORRIGÉ

def edit_product(request, product_id):
    """Éditer un produit individuel avec calcul automatique des prix (TND)"""
    produit = get_object_or_404(Produit, id=product_id)
    
    if request.method == 'POST':
        form = ProduitForm(request.POST, instance=produit)
        if form.is_valid():
            try:
                # 🔥 Récupérer les valeurs du formulaire (AVANT sauvegarde)
                prix = form.cleaned_data.get('prix')
                prix_avant = form.cleaned_data.get('prix_avant')
                pourcentage = form.cleaned_data.get('pourcentage')
                remise = form.cleaned_data.get('remise')
                
                # 🔥 Sauvegarder le formulaire (mais pas encore en DB)
                produit = form.save(commit=False)
                
                # 🔥 Logique de calcul MAIS seulement si les valeurs sont manquantes
                # 1. Si prix et prix_avant sont fournis
                if prix is not None and prix_avant is not None and prix_avant > 0:
                    if prix < prix_avant:
                        # Ne calculer que si pourcentage est None ou 0
                        if pourcentage is None or pourcentage == 0:
                            produit.pourcentage = ((prix_avant - prix) / prix_avant) * 100
                        # Ne calculer que si remise est None ou 0
                        if remise is None or remise == 0:
                            produit.remise = prix_avant - prix
                    else:
                        produit.remise = None
                        produit.pourcentage = None
                
                # 2. Si prix et pourcentage sont fournis (et pas prix_avant)
                elif prix is not None and pourcentage is not None and pourcentage > 0 and pourcentage <= 100:
                    if prix_avant is None:
                        produit.prix_avant = prix / (1 - pourcentage / 100)
                        produit.remise = produit.prix_avant - prix
                
                # 3. Si prix_avant et pourcentage sont fournis (et pas prix)
                elif prix_avant is not None and pourcentage is not None and pourcentage > 0 and pourcentage <= 100:
                    if prix is None:
                        produit.prix = prix_avant * (1 - pourcentage / 100)
                        produit.remise = prix_avant - produit.prix
                
                # 4. Si prix et remise sont fournis
                elif prix is not None and remise is not None and remise > 0:
                    if prix_avant is None:
                        produit.prix_avant = prix + remise
                        produit.pourcentage = (remise / produit.prix_avant) * 100
                
                # Sauvegarder les modifications
                produit.save()
                
                messages.success(request, 'Produit mis à jour avec succès !')
                return redirect('upload')
                
            except Exception as e:
                messages.error(request, f"Erreur lors de la mise à jour: {str(e)}")
        else:
            messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = ProduitForm(instance=produit)
    
    context = {
        'form': form,
        'produit': produit,
    }
    return render(request, 'edit_product.html', context)

@csrf_exempt
def update_product_field(request):
    """API pour mettre à jour un champ d'un produit en AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('product_id')
            field_name = data.get('field_name')
            field_value = data.get('field_value')
            
            if not product_id or not field_name:
                return JsonResponse({'success': False, 'error': 'Parametres manquants'})
            
            produit = get_object_or_404(Produit, id=product_id)
            
            if field_name in ['prix', 'prix_avant', 'pourcentage', 'remise']:
                if field_value:
                    try:
                        setattr(produit, field_name, Decimal(str(field_value)))
                    except:
                        setattr(produit, field_name, None)
                else:
                    setattr(produit, field_name, None)
                
                produit.save()
                
                return JsonResponse({
                    'success': True,
                    'prix': str(produit.prix) if produit.prix else None,
                    'prix_avant': str(produit.prix_avant) if produit.prix_avant else None,
                    'pourcentage': str(produit.pourcentage) if produit.pourcentage else None,
                    'remise': str(produit.remise) if produit.remise else None,
                })
            
            return JsonResponse({'success': False, 'error': 'Champ non modifiable'})
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON invalide'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Methode non autorisee'})

def save_selected_products(request):
    """Sauvegarder les produits sélectionnés (depuis la base de données)"""
    if request.method == 'POST':
        print("=" * 80)
        print("🔍 SAVE SELECTED - AVANT TRAITEMENT")
        print(f"📌 Session avant: {dict(request.session)}")
        print(f"📌 last_uploaded_images avant: {request.session.get('last_uploaded_images', 'NON TROUVÉ')}")
        print(f"📌 active_catalogue_id avant: {request.session.get('active_catalogue_id', 'NON TROUVÉ')}")
        print("=" * 80)
        try:
            # 🔥 Récupérer la note depuis le POST
            note = request.POST.get('note', '').strip()
            catalogue_id = request.session.get('active_catalogue_id', None)
            
            # 🔥 Récupérer le paramètre de redirection
            redirect_to_list = request.POST.get('redirect_to_list', 'false') == 'true'
            
            # 🔥 Sauvegarder la note si elle a changé
            if catalogue_id:
                try:
                    catalogue = Catalogue.objects.get(id=catalogue_id)
                    if catalogue.note != note:
                        catalogue.note = note
                        catalogue.save()
                        print(f"✅ Note mise à jour: '{note}'")
                except Catalogue.DoesNotExist:
                    pass
            
            product_ids = request.POST.getlist('selected_products')
            
            # 🔥 FILTRER : Enlever les IDs vides
            product_ids = [pid for pid in product_ids if pid]
            
            if not product_ids:
                messages.warning(request, 'Aucun produit sélectionné.')
                return redirect('upload')
            
            # 🔥 Sauvegarder UNIQUEMENT les produits sélectionnés non sauvegardés
            produits = Produit.objects.filter(id__in=product_ids, est_sauvegarde=False)
            count = produits.count()
            
            if count == 0:
                messages.warning(request, 'Aucun produit sélectionné non sauvegardé. Ils sont peut-être déjà sauvegardés.')
                return redirect('upload')
            
            produits.update(est_sauvegarde=True)
            
            try:
                produits_a_sync = Produit.objects.filter(id__in=product_ids, est_sauvegarde=True)
                if produits_a_sync.exists():
                    sql_sync.sync_produits(produits_a_sync)
                    print(f"✅ {produits_a_sync.count()} produits synchronisés avec SQL Server")
            except Exception as e:
                print(f"⚠️ Erreur lors de la synchronisation SQL: {e}")

            restants = Produit.objects.filter(est_sauvegarde=False).count()
            
            # 🔥 GARDER le catalogue dans la session si des produits restent
            if restants > 0:
                if 'active_catalogue_id' not in request.session:
                    dernier_catalogue = Catalogue.objects.filter(
                        produits__est_sauvegarde=False
                    ).distinct().order_by('-date_upload').first()
                    if dernier_catalogue:
                        request.session['active_catalogue_id'] = dernier_catalogue.id
                        print(f"✅ Catalogue ID {dernier_catalogue.id} sauvegardé dans la session")
                
                # Garder les images dans la session
                images_actuelles = request.session.get('last_uploaded_images', None)
                if images_actuelles:
                    request.session['last_uploaded_images'] = images_actuelles
                    if images_actuelles:
                        request.session['last_uploaded_image'] = images_actuelles[0]
                    request.session.modified = True
                    print(f"✅ {len(images_actuelles)} images conservées dans la session")
                else:
                    image_actuelle = request.session.get('last_uploaded_image', None)
                    if image_actuelle:
                        request.session['last_uploaded_image'] = image_actuelle
                        print(f"✅ Image conservée dans la session: {image_actuelle}")
            else:
                # 🔥 Plus de produits → Supprimer les données de session
                if 'last_uploaded_images' in request.session:
                    del request.session['last_uploaded_images']
                if 'last_uploaded_image' in request.session:
                    del request.session['last_uploaded_image']
                if 'active_catalogue_id' in request.session:
                    del request.session['active_catalogue_id']
                request.session.modified = True
                request.session.save()
            
            if restants == 0:
                messages.success(request, f'Tous les {count} produits sélectionnés sont sauvegardés !')
            else:
                messages.success(request, f'{count} produits sauvegardés ! Il reste {restants} produit(s) en attente.')
            
            print("=" * 80)
            print("🔍 SAVE SELECTED - APRÈS TRAITEMENT")
            print(f"📌 Session après: {dict(request.session)}")
            print(f"📌 last_uploaded_images après: {request.session.get('last_uploaded_images', 'NON TROUVÉ')}")
            print(f"📌 active_catalogue_id après: {request.session.get('active_catalogue_id', 'NON TROUVÉ')}")
            print(f"📌 restants: {restants}")
            print("=" * 80)
            
            # 🔥 Rediriger vers product_list si demandé
            if redirect_to_list:
                return redirect('upload')
            else:
                return redirect('upload')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la sauvegarde: {str(e)}")
            return redirect('upload')
    return redirect('upload')

def save_all_products(request):
    """Sauvegarder tous les produits et rediriger vers product_list"""
    if request.method == 'POST':
        try:
            # 🔥 Récupérer la note depuis le POST
            note = request.POST.get('note', '').strip()
            catalogue_id = request.session.get('active_catalogue_id', None)
            
            # 🔥 Sauvegarder la note si elle a changé
            if catalogue_id:
                try:
                    catalogue = Catalogue.objects.get(id=catalogue_id)
                    if catalogue.note != note:
                        catalogue.note = note
                        catalogue.save()
                        print(f"✅ Note mise à jour (save_all): '{note}'")
                except Catalogue.DoesNotExist:
                    pass
            
            produits = Produit.objects.filter(est_sauvegarde=False)
            count = produits.count()
            
            if count == 0:
                messages.info(request, 'Aucun nouveau produit a sauvegarder.')
                return redirect('product_list')
            
            produits.update(est_sauvegarde=True)

            try:
                produits_a_sync = Produit.objects.filter(est_sauvegarde=True)
                if produits_a_sync.exists():
                    sql_sync.sync_produits(produits_a_sync)
                    print(f"✅ {produits_a_sync.count()} produits synchronisés avec SQL Server")
            except Exception as e:
                print(f"⚠️ Erreur lors de la synchronisation SQL: {e}")

            # 🔥 Supprimer TOUTES les données de session
            if 'last_uploaded_images' in request.session:
                del request.session['last_uploaded_images']
                print("🗑️ Toutes les images supprimées de la session (save_all)")
            if 'last_uploaded_image' in request.session:
                del request.session['last_uploaded_image']
            if 'active_catalogue_id' in request.session:
                del request.session['active_catalogue_id']
                print("🗑️ active_catalogue_id supprimé (save_all)")
            
            request.session.modified = True
            request.session.save()
            
            messages.success(request, f'{count} produits sauvegardes !')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la sauvegarde: {str(e)}")
        
        return redirect('product_list')
    return redirect('upload')

# catalogue_app/views.py - CORRIGER delete_selected_products
def delete_selected_products(request):
    """Supprimer les produits sélectionnés (depuis la base de données)"""
    if request.method == 'POST':
        print("=" * 80)
        print("🔍 DELETE SELECTED - AVANT TRAITEMENT")
        print(f"📌 Session avant: {dict(request.session)}")
        print(f"📌 active_catalogue_id avant: {request.session.get('active_catalogue_id', 'NON TROUVÉ')}")
        print("=" * 80)
        try:
            # 🔥 Récupérer le paramètre de redirection
            redirect_to_list = request.POST.get('redirect_to_list', 'false') == 'true'
            
            product_ids = request.POST.getlist('selected_products')
            
            # 🔥 FILTRER : Enlever les IDs vides
            product_ids = [pid for pid in product_ids if pid]
            
            if not product_ids:
                messages.warning(request, 'Aucun produit sélectionné.')
                return redirect('upload')
            
            produits = Produit.objects.filter(id__in=product_ids, est_sauvegarde=False)
            count = produits.count()
            
            if count == 0:
                messages.warning(request, 'Aucun produit sélectionné non sauvegardé. Ils sont peut-être déjà supprimés.')
                return redirect('upload')
            
            # 🔥 Récupérer les IDs AVANT suppression
            produits_ids = list(produits.values_list('id', flat=True))
            
            for produit in produits:
                if produit.image_produit:
                    try:
                        image_path = os.path.join(settings.MEDIA_ROOT, str(produit.image_produit))
                        if os.path.exists(image_path):
                            os.remove(image_path)
                    except:
                        pass
            
            produits.delete()
            
            # 🔥 Supprimer de SQL Server
            try:
                if produits_ids:
                    sql_sync.delete_produits(produits_ids)
                    print(f"✅ {len(produits_ids)} produits supprimés de SQL Server")
            except Exception as e:
                print(f"⚠️ Erreur lors de la suppression SQL: {e}")
            
            restants = Produit.objects.filter(est_sauvegarde=False).count()
            
            # 🔥 GARDER le catalogue dans la session si des produits restent
            if restants > 0:
                if 'active_catalogue_id' not in request.session:
                    dernier_catalogue = Catalogue.objects.filter(
                        produits__est_sauvegarde=False
                    ).distinct().order_by('-date_upload').first()
                    if dernier_catalogue:
                        request.session['active_catalogue_id'] = dernier_catalogue.id
                        print(f"✅ Catalogue ID {dernier_catalogue.id} sauvegardé dans la session")
                
                images_actuelles = request.session.get('last_uploaded_images', None)
                if images_actuelles:
                    request.session['last_uploaded_images'] = images_actuelles
                    if images_actuelles:
                        request.session['last_uploaded_image'] = images_actuelles[0]
                    request.session.modified = True
                    print(f"✅ {len(images_actuelles)} images conservées dans la session")
                else:
                    image_actuelle = request.session.get('last_uploaded_image', None)
                    if image_actuelle:
                        request.session['last_uploaded_image'] = image_actuelle
                        print(f"✅ Image conservée dans la session: {image_actuelle}")
            else:
                # 🔥 Plus de produits → Supprimer les données de session
                if 'last_uploaded_images' in request.session:
                    del request.session['last_uploaded_images']
                if 'last_uploaded_image' in request.session:
                    del request.session['last_uploaded_image']
                if 'active_catalogue_id' in request.session:
                    del request.session['active_catalogue_id']
                request.session.modified = True
                request.session.save()
            
            if restants == 0:
                messages.success(request, f'Tous les produits sélectionnés sont supprimés !')
            else:
                messages.success(request, f'{count} produits supprimés ! Il reste {restants} produit(s).')
            
            print("=" * 80)
            print("🔍 DELETE SELECTED - APRÈS TRAITEMENT")
            print(f"📌 Session après: {dict(request.session)}")
            print(f"📌 last_uploaded_images après: {request.session.get('last_uploaded_images', 'NON TROUVÉ')}")
            print(f"📌 active_catalogue_id après: {request.session.get('active_catalogue_id', 'NON TROUVÉ')}")
            print(f"📌 restants: {restants}")
            print("=" * 80)
            
            # 🔥 Rediriger vers product_list si demandé
            if redirect_to_list:
                return redirect('upload')
            else:
                return redirect('upload')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression: {str(e)}")
            return redirect('upload')
    return redirect('upload')
# catalogue_app/views.py - CORRIGER delete_all_products

def delete_all_products(request):
    """Supprimer tous les produits non sauvegardés (depuis la base de données)"""
    if request.method == 'POST':
        try:
            produits = Produit.objects.filter(est_sauvegarde=False)
            count = produits.count()
            
            if count == 0:
                messages.info(request, 'Aucun produit à supprimer.')
                return redirect('upload')
            
            # 🔥 Récupérer les IDs AVANT suppression
            produits_ids = list(produits.values_list('id', flat=True))
            
            # Supprimer les images associées
            for produit in produits:
                if produit.image_produit:
                    try:
                        image_path = os.path.join(settings.MEDIA_ROOT, str(produit.image_produit))
                        if os.path.exists(image_path):
                            os.remove(image_path)
                    except Exception as e:
                        pass
            
            produits.delete()
            
            # 🔥 Supprimer de SQL Server
            try:
                if produits_ids:
                    sql_sync.delete_produits(produits_ids)
                    print(f"✅ {len(produits_ids)} produits supprimés de SQL Server")
            except Exception as e:
                print(f"⚠️ Erreur lors de la suppression SQL: {e}")
            
            # 🔥 Supprimer TOUTES les données de session
            if 'last_uploaded_images' in request.session:
                del request.session['last_uploaded_images']
                print("🗑️ Toutes les images supprimées de la session (delete_all)")
            if 'last_uploaded_image' in request.session:
                del request.session['last_uploaded_image']
            if 'active_catalogue_id' in request.session:
                del request.session['active_catalogue_id']
                print("🗑️ active_catalogue_id supprimé (delete_all)")
            
            request.session.modified = True
            request.session.save()
            
            messages.success(request, f'{count} produits supprimés !')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression: {str(e)}")
        
        return redirect('upload')
    return redirect('upload')

def product_list(request):
    """Liste des produits du dernier catalogue sauvegardé avec filtres"""
    supprimer_catalogues_vides()
    
    # Récupérer les paramètres de filtre
    enseigne_id = request.GET.get('enseigne')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Récupérer tous les catalogues avec produits sauvegardés
    catalogues_avec_produits = []
    for catalogue in Catalogue.objects.order_by('-date_upload'):
        produits_sauvegardes = catalogue.produits.filter(est_sauvegarde=True)
        if produits_sauvegardes.exists():
            catalogues_avec_produits.append({
                'catalogue': catalogue,
                'produits': produits_sauvegardes,
                'count': produits_sauvegardes.count()
            })
    
    # 🔥 APPLIQUER LES FILTRES
    filtered_catalogues = []
    
    for item in catalogues_avec_produits:
        catalogue = item['catalogue']
        match = True
        
        # Filtre par enseigne
        if enseigne_id and enseigne_id != '':
            if str(catalogue.enseigne.id) != enseigne_id:
                match = False
        
        # Filtre par date début
        if date_debut and date_debut != '' and match:
            from datetime import datetime
            try:
                filter_date = datetime.strptime(date_debut, '%Y-%m-%d').date()
                if catalogue.date_debut != filter_date:
                    match = False
            except:
                pass
        
        # Filtre par date fin
        if date_fin and date_fin != '' and match:
            from datetime import datetime
            try:
                filter_date = datetime.strptime(date_fin, '%Y-%m-%d').date()
                if catalogue.date_fin != filter_date:
                    match = False
            except:
                pass
        
        if match:
            filtered_catalogues.append(item)
    
    # 🔥 Si des filtres sont appliqués, afficher TOUS les catalogues filtrés
    if enseigne_id or date_debut or date_fin:
        catalogues_a_afficher = filtered_catalogues
        has_filters = True
    else:
        # Pas de filtre : prendre seulement le dernier catalogue
        catalogues_a_afficher = catalogues_avec_produits[:1]  # Seulement le premier
        has_filters = False
    
    # 🔥 Récupérer les données pour le template
    if catalogues_a_afficher:
        # Pour l'affichage principal, on prend le premier catalogue
        premier_item = catalogues_a_afficher[0]
        dernier_catalogue = premier_item['catalogue']
        produits = premier_item['produits']
        
        # 🔥 Tous les catalogues pour l'affichage des filtres
        tous_les_catalogues = catalogues_a_afficher
    else:
        dernier_catalogue = None
        produits = []
        tous_les_catalogues = []
    
    # Récupérer toutes les enseignes pour le filtre
    enseignes = Enseigne.objects.all().order_by('nom')
    
    context = {
        'catalogue': dernier_catalogue,           # Catalogue principal affiché
        'produits': produits,                     # Produits du catalogue principal
        'tous_les_catalogues': tous_les_catalogues,  # 🔥 TOUS les catalogues filtrés
        'total_produits': produits.count() if produits else 0,
        'has_products': bool(produits),
        'enseignes': enseignes,
        'filtre_enseigne': enseigne_id,
        'filtre_date_debut': date_debut,
        'filtre_date_fin': date_fin,
        'has_filters': has_filters,
        'nombre_catalogues': len(tous_les_catalogues),
    }
    
    return render(request, 'product_list.html', context)

def get_product_details(request, product_id):
    """API pour récupérer les détails d'un produit (AJAX)"""
    if request.method == 'GET':
        try:
            produit = get_object_or_404(Produit, id=product_id)
            data = {
                'id': produit.id,
                'nom': produit.nom,
                'prix': str(produit.prix) if produit.prix else None,
                'prix_avant': str(produit.prix_avant) if produit.prix_avant else None,
                'pourcentage': str(produit.pourcentage) if produit.pourcentage else None,
                'remise': str(produit.remise) if produit.remise else None,
                'desc_1': produit.desc_1,
                'extrait_texte': produit.extrait_texte,
                'est_sauvegarde': produit.est_sauvegarde,
                'image_url': produit.image_produit.url if produit.image_produit else None,
            }
            return JsonResponse({'success': True, 'data': data})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Methode non autorisee'})


def get_catalogue_image(request, catalogue_id):
    """API pour récupérer l'image du catalogue"""
    if request.method == 'GET':
        try:
            catalogue = get_object_or_404(Catalogue, id=catalogue_id)
            if catalogue.image_path:
                image_url = os.path.join(settings.MEDIA_URL, catalogue.image_path)
                return JsonResponse({'success': True, 'image_url': image_url})
            return JsonResponse({'success': False, 'error': 'Aucune image'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Methode non autorisee'})


def get_recent_products(request):
    """Récupère les produits non sauvegardés récents"""
    produits = Produit.objects.filter(est_sauvegarde=False).order_by('-created_at')[:10]
    data = []
    for p in produits:
        data.append({
            'id': p.id,
            'nom': p.nom,
            'nom_fr': p.nom_fr,
            'nom_ar': p.nom_ar,
            'marque': p.marque,
            'prix': str(p.prix) if p.prix else None,
            'prix_avant': str(p.prix_avant) if p.prix_avant else None,
            'pourcentage': str(p.pourcentage) if p.pourcentage else None,
            'desc_1': p.desc_1,
            'desc_2': p.desc_2,
            'desc_3': p.desc_3,
            'image_url': p.image_produit.url if p.image_produit else None,
        })
    return JsonResponse({'products': data})
def add_product(request):
    """Page pour ajouter un produit manuellement (sans extraction)"""
    
    # 🔥 Récupérer le catalogue depuis la session
    catalogue_id = request.session.get('active_catalogue_id')
    catalogue = None
    
    # 🔥 Si un catalogue existe dans la session
    if catalogue_id:
        try:
            catalogue = Catalogue.objects.get(id=catalogue_id)
        except Catalogue.DoesNotExist:
            catalogue = None
    
    # 🔥 SI PAS DE CATALOGUE EN SESSION, CHERCHER DANS LA BASE
    if not catalogue:
        # Chercher le dernier catalogue avec des produits non sauvegardés
        dernier_produit = Produit.objects.filter(est_sauvegarde=False).order_by('-created_at').first()
        if dernier_produit and dernier_produit.catalogue:
            catalogue = dernier_produit.catalogue
            # 🔥 Important : Sauvegarder dans la session pour la prochaine fois
            request.session['active_catalogue_id'] = catalogue.id
            print(f"✅ Catalogue récupéré depuis la base: ID {catalogue.id}")
    
    # 🔥 SI TOUJOURS PAS DE CATALOGUE, CRÉER UN CATALOGUE TEMPORAIRE
    if not catalogue:
        try:
            from datetime import datetime, timedelta
            from django.utils import timezone
            
            enseigne = Enseigne.objects.first()
            if not enseigne:
                enseigne = Enseigne.objects.create(nom='autre')
            
            today = datetime.now().date()
            catalogue = Catalogue.objects.create(
                enseigne=enseigne,
                date_debut=today,
                date_fin=today + timedelta(days=7),
                date_upload=timezone.now(),
                note="Catalogue temporaire"
            )
            request.session['active_catalogue_id'] = catalogue.id
            print(f"✅ Catalogue temporaire créé: ID {catalogue.id}")
            
        except Exception as e:
            print(f"❌ Erreur création catalogue: {e}")
            messages.error(request, "❌ Aucun catalogue disponible. Veuillez d'abord uploader une image.")
            return redirect('upload')
    
    # 🔥 Si toujours pas de catalogue, rediriger
    if not catalogue:
        messages.error(request, "❌ Aucun catalogue actif. Veuillez d'abord uploader une image.")
        return redirect('upload')
    
    if request.method == 'POST':
        # Récupérer les données du formulaire
        nom_fr = request.POST.get('nom_fr', '').strip()
        nom_ar = request.POST.get('nom_ar', '').strip()
        marque = request.POST.get('marque', '').strip()
        desc_1 = request.POST.get('desc_1', '').strip()
        desc_2 = request.POST.get('desc_2', '').strip()
        desc_3 = request.POST.get('desc_3', '').strip()
        note_1 = request.POST.get('note_1', '').strip()
        note_2 = request.POST.get('note_2', '').strip()
        
        # Convertir les prix
        prix = None
        if request.POST.get('prix'):
            try:
                prix = Decimal(request.POST.get('prix').replace(',', '.'))
            except:
                pass
        
        prix_avant = None
        if request.POST.get('prix_avant'):
            try:
                prix_avant = Decimal(request.POST.get('prix_avant').replace(',', '.'))
            except:
                pass
        
        pourcentage = None
        if request.POST.get('pourcentage'):
            try:
                pourcentage = Decimal(request.POST.get('pourcentage').replace(',', '.'))
            except:
                pass

        # Gérer l'image du produit
        image_produit = None
        if 'product_image' in request.FILES:
            image_produit = request.FILES['product_image']
                
        # Validation - au moins un nom
        if not nom_fr and not nom_ar:
            messages.error(request, "❌ Veuillez saisir au moins un nom (FR ou AR).")
            return render(request, 'add_product.html', {'catalogue': catalogue})
        
        # Créer le produit
        produit = Produit(
            catalogue=catalogue,
            nom=nom_fr or nom_ar,
            nom_fr=nom_fr or None,
            nom_ar=nom_ar or None,
            marque=marque or None,
            prix=prix,
            prix_avant=prix_avant,
            pourcentage=pourcentage,
            desc_1=desc_1 or None,
            desc_2=desc_2 or None,
            desc_3=desc_3 or None,
            note_1=note_1 or None, 
            note_2=note_2 or None,
            image_produit=image_produit,
            est_sauvegarde=False,
        )
        produit.save()
        
        messages.success(request, f"✅ Produit '{produit.nom}' ajouté avec succès !")
        return redirect('upload')
    
    context = {
        'catalogue': catalogue,
    }
    return render(request, 'add_product.html', context)

def get_marques_list(request):
    """API pour récupérer la liste des marques uniques des produits non sauvegardés"""
    if request.method == 'GET':
        try:
            # Récupérer toutes les marques uniques des produits non sauvegardés
            marques = Produit.objects.filter(
                est_sauvegarde=False
            ).exclude(
                marque__isnull=True
            ).exclude(
                marque=''
            ).values_list('marque', flat=True).distinct().order_by('marque')
            
            # Filtrer les marques vides ou None
            marques_list = [m for m in marques if m and m.strip()]
            
            print(f"🔍 Marques trouvées: {marques_list}")
            
            return JsonResponse({
                'success': True,
                'marques': marques_list
            })
        except Exception as e:
            print(f"❌ Erreur get_marques_list: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

def edit_product_saved(request, product_id):
    """Éditer un produit sauvegardé (redirige vers product_list après sauvegarde)"""
    produit = get_object_or_404(Produit, id=product_id)
    
    if request.method == 'POST':
        form = ProduitForm(request.POST, instance=produit)
        if form.is_valid():
            try:
                # 🔥 Récupérer les valeurs du formulaire (AVANT sauvegarde)
                prix = form.cleaned_data.get('prix')
                prix_avant = form.cleaned_data.get('prix_avant')
                pourcentage = form.cleaned_data.get('pourcentage')
                remise = form.cleaned_data.get('remise')
                
                # 🔥 Sauvegarder le formulaire (mais pas encore en DB)
                produit = form.save(commit=False)
                
                # 🔥 Logique de calcul MAIS seulement si les valeurs sont manquantes
                # 1. Si prix et prix_avant sont fournis
                if prix is not None and prix_avant is not None and prix_avant > 0:
                    if prix < prix_avant:
                        # Ne calculer que si pourcentage est None ou 0
                        if pourcentage is None or pourcentage == 0:
                            produit.pourcentage = ((prix_avant - prix) / prix_avant) * 100
                        # Ne calculer que si remise est None ou 0
                        if remise is None or remise == 0:
                            produit.remise = prix_avant - prix
                    else:
                        produit.remise = None
                        produit.pourcentage = None
                
                # 2. Si prix et pourcentage sont fournis (et pas prix_avant)
                elif prix is not None and pourcentage is not None and pourcentage > 0 and pourcentage <= 100:
                    if prix_avant is None:
                        produit.prix_avant = prix / (1 - pourcentage / 100)
                        produit.remise = produit.prix_avant - prix
                
                # 3. Si prix_avant et pourcentage sont fournis (et pas prix)
                elif prix_avant is not None and pourcentage is not None and pourcentage > 0 and pourcentage <= 100:
                    if prix is None:
                        produit.prix = prix_avant * (1 - pourcentage / 100)
                        produit.remise = prix_avant - produit.prix
                
                # 4. Si prix et remise sont fournis
                elif prix is not None and remise is not None and remise > 0:
                    if prix_avant is None:
                        produit.prix_avant = prix + remise
                        produit.pourcentage = (remise / produit.prix_avant) * 100
                
                # 🔥 Sauvegarder les modifications
                produit.save()
                
                # 🔥🔥🔥 AJOUT : Synchronisation SQL Server
                try:
                    sql_sync.sync_produits([produit])
                    print(f"✅ Produit {produit.id} synchronisé avec SQL Server")
                except Exception as e:
                    print(f"⚠️ Erreur lors de la synchronisation SQL: {e}")
                
                messages.success(request, 'Produit mis à jour avec succès !')
                return redirect('product_list')  # 🔥 SEULE DIFFÉRENCE : redirection vers product_list
                
            except Exception as e:
                messages.error(request, f"Erreur lors de la mise à jour: {str(e)}")
        else:
            messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = ProduitForm(instance=produit)
    
    context = {
        'form': form,
        'produit': produit,
    }
    return render(request, 'edit_product_saved.html', context)

@csrf_exempt
def delete_saved_product(request, product_id):
    """Supprimer un produit sauvegardé"""
    if request.method == 'POST':
        try:
            produit = get_object_or_404(Produit, id=product_id)
            
            # 🔥 Récupérer l'ID avant suppression
            produit_id = produit.id
            print(f"🗑️ Suppression du produit ID: {produit_id}")
            
            # Supprimer l'image associée
            if produit.image_produit:
                try:
                    image_path = os.path.join(settings.MEDIA_ROOT, str(produit.image_produit))
                    if os.path.exists(image_path):
                        os.remove(image_path)
                        print(f"✅ Image supprimée: {image_path}")
                except Exception as e:
                    print(f"⚠️ Erreur suppression image: {e}")
            
            produit.delete()
            print(f"✅ Produit {produit_id} supprimé de SQLite")
            
            # 🔥🔥🔥 AJOUT : Supprimer de SQL Server
            try:
                result = sql_sync.delete_produits([produit_id])
                if result:
                    print(f"✅ Produit {produit_id} supprimé de SQL Server")
                else:
                    print(f"⚠️ Échec suppression SQL Server pour {produit_id}")
            except Exception as e:
                print(f"⚠️ Erreur lors de la suppression SQL: {e}")
            
            return JsonResponse({'success': True, 'message': 'Produit supprimé avec succès'})
            
        except Exception as e:
            print(f"❌ Erreur delete_saved_product: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

def export_products_excel(request):
    """Exporter les produits du dernier catalogue sauvegardé en fichier Excel"""
    
    # 🔥 Récupérer le dernier catalogue QUI A DES PRODUITS SAUVEGARDÉS
    dernier_catalogue = None
    
    for catalogue in Catalogue.objects.order_by('-date_upload'):
        produits_sauvegardes = catalogue.produits.filter(est_sauvegarde=True)
        if produits_sauvegardes.exists():
            dernier_catalogue = catalogue
            produits = produits_sauvegardes
            break
    
    if not dernier_catalogue:
        messages.warning(request, "Aucun catalogue sauvegardé à exporter.")
        return redirect('product_list')
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Produits - {dernier_catalogue.enseigne.nom}"
    
    # 🔥 En-têtes avec les informations du catalogue
    headers = [
        "#", "Nom (FR)", "Nom (AR)", "Marque", 
        "Prix (DT)", "Prix avant (DT)", "Pourcentage (%)", "Remise (DT)",
        "Desc_1", "Desc_2", "Desc_3",
        "Note 1", "Note 2"
    ]
    
    # 🔥 Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="0073E6", end_color="0073E6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 🔥 Écrire les données
    for row_idx, produit in enumerate(produits, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Numéro de ligne
        ws.cell(row=row_idx, column=2, value=produit.nom_fr or '')
        ws.cell(row=row_idx, column=3, value=produit.nom_ar or '')
        ws.cell(row=row_idx, column=4, value=produit.marque or '')
        ws.cell(row=row_idx, column=5, value=float(produit.prix) if produit.prix else None)
        ws.cell(row=row_idx, column=6, value=float(produit.prix_avant) if produit.prix_avant else None)
        ws.cell(row=row_idx, column=7, value=float(produit.pourcentage) if produit.pourcentage else None)
        ws.cell(row=row_idx, column=8, value=float(produit.remise) if produit.remise else None)
        ws.cell(row=row_idx, column=9, value=produit.desc_1 or '')
        ws.cell(row=row_idx, column=10, value=produit.desc_2 or '')
        ws.cell(row=row_idx, column=11, value=produit.desc_3 or '')
        ws.cell(row=row_idx, column=12, value=produit.note_1 or '')
        ws.cell(row=row_idx, column=13, value=produit.note_2 or '')
        
        # 🔥 Appliquer le style aux cellules
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if col == 5 or col == 6 or col == 8:  # Prix, Prix avant, Remise
                cell.number_format = '#,##0.000'
                cell.alignment = Alignment(horizontal="right")
            elif col == 7:  # 🔥 Pourcentage - format sans décimales
                cell.number_format = '0'  # Affiche le nombre sans décimales
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
    
    # 🔥 Ajuster la largeur des colonnes
    column_widths = {
        'A': 8, 'B': 30, 'C': 30, 'D': 25,
        'E': 15, 'F': 15, 'G': 15, 'H': 15,
        'I': 35, 'J': 35, 'K': 35, 'L': 35, 'M': 35
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 🔥 Ajouter une feuille avec les informations du catalogue
    ws_info = wb.create_sheet("Informations catalogue")
    info_data = [
        ["Enseigne", dernier_catalogue.enseigne.nom if dernier_catalogue.enseigne else ""],
        ["Date début", dernier_catalogue.date_debut.strftime('%d/%m/%Y') if dernier_catalogue.date_debut else ""],
        ["Date fin", dernier_catalogue.date_fin.strftime('%d/%m/%Y') if dernier_catalogue.date_fin else ""],
        ["Date upload", dernier_catalogue.date_upload.strftime('%d/%m/%Y %H:%M') if dernier_catalogue.date_upload else ""],
        ["Nombre de produits", produits.count()],
    ]
    
    for row_idx, (label, value) in enumerate(info_data, 1):
        ws_info.cell(row=row_idx, column=1, value=label)
        ws_info.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_info.cell(row=row_idx, column=2, value=value)
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 30
    
    # 🔥 Ligne de total sur la feuille principale
    total_row = len(produits) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{total_row-1})")
    
    # 🔥 Créer le nom du fichier avec les infos du catalogue
    filename = f"produits_{dernier_catalogue.enseigne.nom}_{dernier_catalogue.date_debut.strftime('%Y%m%d')}.xlsx"
    
    # 🔥 Créer la réponse HTTP avec le fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# catalogue_app/views.py - Ajouter cette fonction après export_products_excel
# catalogue_app/views.py - Version améliorée de export_catalogue_excel

def export_catalogue_excel(request, catalogue_id):
    """Exporter les produits d'un catalogue spécifique en fichier Excel sans répétition des infos du catalogue"""
    
    # 🔥 Récupérer le catalogue
    catalogue = get_object_or_404(Catalogue, id=catalogue_id)
    produits = catalogue.produits.filter(est_sauvegarde=True)
    
    if not produits.exists():
        messages.warning(request, "Aucun produit sauvegardé dans ce catalogue à exporter.")
        return redirect('product_list')
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Produits - {catalogue.enseigne.nom}"
    
    # 🔥 LIGNE 1 : TITRE AVEC INFOS DU CATALOGUE (fusionnée)
    titre = f"Catalogue: {catalogue.enseigne.nom} - Du {catalogue.date_debut.strftime('%d/%m/%Y')} au {catalogue.date_fin.strftime('%d/%m/%Y')}"
    if catalogue.note:
        titre += f" - {catalogue.note}"
    
    ws.merge_cells('A1:N1')
    cell_titre = ws.cell(row=1, column=1, value=titre)
    cell_titre.font = Font(bold=True, size=14, color="FFFFFF")
    cell_titre.fill = openpyxl.styles.PatternFill(start_color="0073E6", end_color="0073E6", fill_type="solid")
    cell_titre.alignment = Alignment(horizontal="center", vertical="center")
    
    # 🔥 LIGNE 2 : SOUS-TITRE AVEC LE NOMBRE DE PRODUITS
    ws.merge_cells('A2:N2')
    cell_sous_titre = ws.cell(row=2, column=1, value=f"Nombre de produits: {produits.count()}")
    cell_sous_titre.font = Font(bold=True, size=11)
    cell_sous_titre.alignment = Alignment(horizontal="center", vertical="center")
    cell_sous_titre.fill = openpyxl.styles.PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    
    # 🔥 LIGNE 3 : ESPACE
    ws.row_dimensions[3].height = 5
    
    # 🔥 LIGNE 4 : EN-TÊTES DES COLONNES (SANS Image, Enseigne, Date début, Date fin)
    headers = [
        "#", "Nom (FR)", "Nom (AR)", "Marque", 
        "Prix (DT)", "Prix avant (DT)", "Pourcentage (%)", "Remise (DT)",
        "Desc_1", "Desc_2", "Desc_3",
        "Note 1", "Note 2"
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = openpyxl.styles.PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 🔥 Écrire les données (commence à la ligne 5)
    for row_idx, produit in enumerate(produits, 5):
        # Numéro de ligne
        ws.cell(row=row_idx, column=1, value=row_idx - 4)
        ws.cell(row=row_idx, column=2, value=produit.nom_fr or '')
        ws.cell(row=row_idx, column=3, value=produit.nom_ar or '')
        ws.cell(row=row_idx, column=4, value=produit.marque or '')
        ws.cell(row=row_idx, column=5, value=float(produit.prix) if produit.prix else None)
        ws.cell(row=row_idx, column=6, value=float(produit.prix_avant) if produit.prix_avant else None)
        ws.cell(row=row_idx, column=7, value=float(produit.pourcentage) if produit.pourcentage else None)
        ws.cell(row=row_idx, column=8, value=float(produit.remise) if produit.remise else None)
        ws.cell(row=row_idx, column=9, value=produit.desc_1 or '')
        ws.cell(row=row_idx, column=10, value=produit.desc_2 or '')
        ws.cell(row=row_idx, column=11, value=produit.desc_3 or '')
        ws.cell(row=row_idx, column=12, value=produit.note_1 or '')
        ws.cell(row=row_idx, column=13, value=produit.note_2 or '')
        
        # 🔥 Appliquer le style aux cellules
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if col == 5 or col == 6 or col == 8:  # Prix, Prix avant, Remise
                cell.number_format = '#,##0.000'
                cell.alignment = Alignment(horizontal="right")
            elif col == 7:  # Pourcentage
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
    
    # 🔥 LIGNE DE TOTAL
    total_row = len(produits) + 5
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=1).fill = openpyxl.styles.PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Fusionner les colonnes pour le label TOTAL
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Formules de total
    for col in [5, 6, 8]:
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).number_format = '#,##0.000'
        ws.cell(row=total_row, column=col).fill = openpyxl.styles.PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        ws.cell(row=total_row, column=col).border = border
    
    # 🔥 Ajuster la largeur des colonnes
    column_widths = {
        'A': 8,   # #
        'B': 35,  # Nom FR
        'C': 35,  # Nom AR
        'D': 25,  # Marque
        'E': 18,  # Prix
        'F': 18,  # Prix avant
        'G': 15,  # Pourcentage
        'H': 18,  # Remise
        'I': 40,  # Desc 1
        'J': 40,  # Desc 2
        'K': 40,  # Desc 3
        'L': 30,  # Note 1
        'M': 30,  # Note 2
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 🔥 Ajouter une 2ème feuille avec les informations détaillées du catalogue
    ws_info = wb.create_sheet("Informations catalogue")
    
    # Style pour les infos
    info_title_font = Font(bold=True, size=14, color="FFFFFF")
    info_title_fill = openpyxl.styles.PatternFill(start_color="0073E6", end_color="0073E6", fill_type="solid")
    info_label_font = Font(bold=True, size=11)
    info_value_font = Font(size=11)
    
    # Titre de la feuille d'info
    ws_info.merge_cells('A1:D1')
    cell_title = ws_info.cell(row=1, column=1, value="INFORMATIONS DU CATALOGUE")
    cell_title.font = info_title_font
    cell_title.fill = info_title_fill
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données
    info_data = [
        ["Enseigne", catalogue.enseigne.nom if catalogue.enseigne else ""],
        ["Date début", catalogue.date_debut.strftime('%d/%m/%Y') if catalogue.date_debut else ""],
        ["Date fin", catalogue.date_fin.strftime('%d/%m/%Y') if catalogue.date_fin else ""],
        ["Date upload", catalogue.date_upload.strftime('%d/%m/%Y %H:%M') if catalogue.date_upload else ""],
        ["Note", catalogue.note or ""],
        ["Nombre de produits", produits.count()],
        ["Date d'export", timezone.now().strftime('%d/%m/%Y %H:%M')],
    ]
    
    border_info = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx, (label, value) in enumerate(info_data, 3):
        # Cellule label
        cell_label = ws_info.cell(row=row_idx, column=1, value=label)
        cell_label.font = info_label_font
        cell_label.alignment = Alignment(horizontal="left", vertical="center")
        cell_label.border = border_info
        cell_label.fill = openpyxl.styles.PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        
        # Cellule valeur
        cell_value = ws_info.cell(row=row_idx, column=2, value=value)
        cell_value.font = info_value_font
        cell_value.alignment = Alignment(horizontal="left", vertical="center")
        cell_value.border = border_info
        
        # Fusionner les colonnes C et D pour plus d'espace
        ws_info.merge_cells(f'C{row_idx}:D{row_idx}')
    
    # Ajuster les colonnes
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 40
    ws_info.column_dimensions['C'].width = 10
    ws_info.column_dimensions['D'].width = 10
    
    # 🔥 Créer le nom du fichier
    filename = f"produits_{catalogue.enseigne.nom}_{catalogue.date_debut.strftime('%Y%m%d')}.xlsx"
    
    # 🔥 Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response